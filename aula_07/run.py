
import os
import csv
import openpyxl

arquivo = "logs.log"
pasta = "../aula_07"

# if os.path.exists(arquivo): print("Arquivo existe!")

# if os.path.isdir(pasta): print("Tem esta pasta!")

# print(  os.getcwd() ) # Mostra o caminho da pasta
# os.chdir("../") # Navega até o caminho (Relativo ou absoluto)

# os.mkdir("nova pasta") # Cria pasta

# os.rmdir("nova pasta") # Remove pasta

# os.system("rm arquivo_criado.txt") # comandos no terminal

# print( os.listdir() )


# CSV
alunos_novos = [
    ["Julio Pereira", "julio@email"],
    ["Ana Clara", "ana@gmail.com"]
]

# arq_csv = open("alunos_novos.csv", "w")

# gravar = csv.writer(arq_csv)

# gravar.writerow(["NOME", "EMAIL"])
# gravar.writerows(alunos_novos)

# arq_csv.close()


## FORMA PARA NÃO PRECISAR DE FECHAR.
# with open("alunos_novos.csv", "w") as arq_csv:
#     gravar = csv.writer(arq_csv)
#     gravar.writerow(["NOME", "EMAIL"])
#     gravar.writerows(alunos_novos)


# EXCEL
pasta = openpyxl.load_workbook(filename="alunos.xlsx")

dados = []

for x in pasta['plan1'].iter_rows(values_only=True):
    dados.append(x)

# print(dados)

pasta = openpyxl.Workbook()
ws = pasta.active

ws["A1"] = "ALUNO"
ws["B1"] = "TURMA"
ws.append(["Julio", "A"])
ws.append(["Cezar", "B"])

pasta.save("alunos_02.xlsx")


# EXCEL (PANDAS)
import pandas as pd

dados = {
    "ALUNO": ["JULIO", "CEZAR"],
    "TURMA": [1, 2]
}

# DATAFRAME | SERIES
df = pd.DataFrame(dados)

# print(df.shape) #mostra quantidade linha e coluna
# print(df.head()) # mostra as 05 primeiras linhas
# print(df.info()) # Mostra informações
# print(df.describe()) # Mostra descrições

# df.plot()

df.to_excel("alunos_02_panda.xlsx", index=False)

excel = pd.read_excel("alunos.xlsx")

print(excel.head())








